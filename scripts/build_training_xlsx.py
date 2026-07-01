"""Build the training-phrase quality Excel from the scorer CSV.

Usage: python scripts/build_training_xlsx.py [lang] [output.xlsx]
Reads app/build/training-phrase-quality-<lang>.csv (produced by TrainingPhraseQualityProbe).
Default lang: ru. Supported shortcut: python scripts/build_training_xlsx.py en output.xlsx
Sheets: data (per-command diagnosis), Legend (column meanings), Method (scoring method + examples).
"""
import csv
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SUPPORTED_LANGS = {"ru", "en"}

args = sys.argv[1:]
LANG = "ru"
if args and args[0].lower() in SUPPORTED_LANGS:
    LANG = args.pop(0).lower()

SRC = fr"app/build/training-phrase-quality-{LANG}.csv"
DST = args[0] if args else fr"C:\Users\Alex\Downloads\training-phrase-quality-{LANG}.xlsx"

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF")
WEAK_FILL = PatternFill("solid", fgColor="F4CCCC")
WATCH_FILL = PatternFill("solid", fgColor="FFF2CC")
TOP_WRAP = Alignment(vertical="top", wrap_text=True)

with open(SRC, encoding="utf-8") as f:
    rows = list(csv.reader(f))
header, data = rows[0], rows[1:]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "data"
ws.append(header)
for c in ws[1]:
    c.fill = HDR_FILL
    c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
vi = header.index("verdict")
for r in data:
    ws.append(r)
    row = ws.max_row
    for c in ws[row]:
        c.alignment = TOP_WRAP
    fill = WEAK_FILL if r[vi] == "WEAK" else WATCH_FILL if r[vi] == "WATCH" else None
    if fill:
        for c in ws[row]:
            c.fill = fill
    lines = max((str(v).count("\n") + 1 for v in r), default=1)
    ws.row_dimensions[row].height = min(15 * lines, 300)
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"
widths = {"id": 28, "category": 9, "verdict": 8, "score": 20, "conflict_group": 26, "probe": 34, "rank": 6,
          "own_match": 30, "competitor": 42, "existing_phrases": 26, "suggested_additions": 32}
for i, h in enumerate(header, 1):
    ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 18)


def sheet(title, pairs, w0, w1, height):
    sh = wb.create_sheet(title)
    for pair in pairs:
        sh.append(pair)
    for c in sh[1]:
        c.fill = HDR_FILL
        c.font = HDR_FONT
    sh.column_dimensions["A"].width = w0
    sh.column_dimensions["B"].width = w1
    for i in range(2, sh.max_row + 1):
        sh.cell(i, 2).alignment = Alignment(wrap_text=True, vertical="top")
        sh.cell(i, 1).alignment = Alignment(wrap_text=True, vertical="top")
        sh.row_dimensions[i].height = height
    return sh


sheet("Legend", [
    ("Column", "Meaning"),
    ("rank", "The target command's position among all ~188 commands for one semantic probe. rank 1 = top match. rank above 1 = other commands scored higher."),
    ("score", "Aggregate over all 10 probes. hit@1 X/10 = how many probes ranked the command first. offered Y/10 = how many probes included the command in the shortlist shown to the model, even when it was not first."),
    ("verdict", "OK = the command was offered for all 10 probes. WATCH = missed once (9/10). WEAK = missed on 2+ probes (<=8/10), so the model will not see the command for those phrasings. The LLM shortlist is capped at 8 game commands plus system functions."),
    ("conflict_group", "The command that most often outranked this one, with the number of occurrences. Shows the main semantic conflict. A dash means no conflict."),
    ("probe", "The 10 test phrases, written as realistic spoken requests, that were searched semantically."),
    ("own_match", "The best matching training phrase from this command for the probe, plus its similarity score (0..1)."),
    ("competitor", "When the command is not rank 1: the competing command and phrase that outranked it, plus similarity. The gap is competitor minus own_match."),
    ("existing_phrases", "Current training phrases for the command after placeholder normalization for embedding."),
    ("suggested_additions", "Failed probes that are candidates for new or improved training phrases after human review."),
], 20, 110, 45)

sheet("Method", [
    ("Step", "Description and example"),
    ("Purpose", "Find commands whose training phrases are weak: realistic player wording does not bring the correct command into the candidate shortlist. The test is deterministic and uses only the embedding model, not a live LLM."),
    ("1. Embeddings", "Each phrase is converted into a meaning vector with multilingual-e5. Semantically close phrases have close vectors. Similarity is cosine distance (0..1): 1 means the same meaning, near 0 means unrelated."),
    ("2. Command score", "For one probe, every command is scored against all of its training phrases. The command score is the best similarity among its phrases (max)."),
    ("3. rank", "All ~188 commands are sorted by score. The target command's position is rank. Example: for probe 'target the engines', target_subsystem may score 0.89 from 'target engines', but query_ship_loadout might score 0.92 from a broad 'engines' phrase, putting target_subsystem at rank 5."),
    ("4. Shortlist (offered)", "The model does not see all commands. The shortlist starts from the best score; if it is below 0.80, the shortlist is empty. Otherwise, keep commands within 0.04 of the best score, capped at 8 commands. offered means the target command made that shortlist."),
    ("5. Gap", "For a failed probe, gap = competitor score minus target command score. Small gaps are often fixed by one good training phrase."),
    ("6. Command summary", "hit@1 = number of probes where the command ranked first. offered = number of probes where the command reached the shortlist. WEAK means offered < 8/10."),
    ("7. What to add", "A failed probe is a candidate training phrase after review. Adding it verbatim usually makes that probe rank 1 because the probe matches itself, but prefer canonical, natural aliases rather than overfitting every probe."),
    ("Important", "The winning similarity almost always comes from a training phrase in the same language as the input. The English command description is weaker for localized input, so quality depends on localized training phrases."),
], 26, 120, 70)

wb.save(DST)
print("wrote", DST, "data rows =", len(data))
